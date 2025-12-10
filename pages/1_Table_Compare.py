import streamlit as st
import pandas as pd
import io
import datetime
import os
import traceback
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


# STREAMLIT CONFIG =========================

st.set_page_config(
    page_title="Table Comparison Tool",
    page_icon="📊",
    layout="wide",
)
st.title("📊 Table Comparison Tool")
st.markdown("Utility to compare two tabular files and review their differences.")


# FUNCTIONS =========================

def read_file_to_df(uploaded_or_path):
    """Read an uploaded file or a file path into a cleaned pandas DataFrame."""
    try:
        if isinstance(uploaded_or_path, str):
            path = uploaded_or_path
            ext = os.path.splitext(path)[1].lower()
            if ext in [".csv", ".txt"]:
                df = pd.read_csv(path, sep=None, engine="python", dtype=str)
            else:
                df = pd.read_excel(path, sheet_name=0, dtype=str)
        else:
            uploaded = uploaded_or_path
            name = getattr(uploaded, "name", "") or ""
            ext = os.path.splitext(name)[1].lower()
            uploaded_file_buffer = uploaded.getvalue()

            if ext in [".csv", ".txt"]:
                try:
                    df = pd.read_csv(
                        io.BytesIO(uploaded_file_buffer),
                        sep=None,
                        engine="python",
                        dtype=str,
                    )
                except Exception:
                    df = pd.read_csv(
                        io.BytesIO(uploaded_file_buffer),
                        dtype=str,
                        engine="python",
                        sep=",",
                        on_bad_lines="skip",
                    )
            else:
                df = pd.read_excel(
                    io.BytesIO(uploaded_file_buffer),
                    sheet_name=0,
                    dtype=str,
                )

        df = df.fillna("").astype(str).apply(lambda x: x.str.strip())
        df = df.replace({"nan": "", "NaN": "", "None": ""})
        df.columns = df.columns.str.strip().str.replace(r"[\r\n]+", "", regex=True)
        return df

    except Exception as e:
        st.error(f"Error while reading file: {e}")
        st.exception(e)
        raise


def apply_excel_coloring(df_export, file_path):

    try:
        if df_export.empty:
            st.warning("DataFrame is empty. Nothing to color.")
            df_export.to_excel(file_path, index=False)
            return file_path

        df_color_logic = df_export.copy()
        df_visible = df_export.drop(columns=["_changed_cols"], errors="ignore")

        df_visible.to_excel(file_path, index=False)
        wb = load_workbook(file_path)
        ws = wb.active

        fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        fill_yellow_light = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        fill_yellow_bright = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        headers = [cell.value for cell in ws[1]]
        header_to_idx = {h: i + 1 for i, h in enumerate(headers)}

        for r_idx, row in df_color_logic.iterrows():
            r = r_idx + 2
            status = row["Status"]
            changed_cols_str = row.get("_changed_cols", "") or ""
            changed_cols = [c for c in changed_cols_str.split(",") if c]

            if status == "Added":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_green

            elif status == "Deleted":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_red

            elif status in ["Duplicate", "MergedDuplicate"]:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_blue

            elif status == "Modified":
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = fill_yellow_light

            if status == "Modified" and changed_cols:
                for colname in changed_cols:
                    if colname in header_to_idx:
                        col_idx = header_to_idx[colname]
                        ws.cell(row=r, column=col_idx).fill = fill_yellow_bright

        wb.save(file_path)
        return file_path

    except Exception as e:
        st.error(f"Error while applying Excel styling: {e}")
        st.exception(e)
        return file_path

def get_compare_summary(df):
    """Return counts by status for summary badges."""
    return {
        "All": int((df["Status"] != "Same").sum()),
        "Added": int((df["Status"] == "Added").sum()),
        "Modified": int((df["Status"] == "Modified").sum()),
        "Deleted": int((df["Status"] == "Deleted").sum()),
        "Duplicate": int((df["Status"] == "Duplicate").sum()),
    }

def cleanup_temp_files(*files):
    """Remove temporary files if they exist."""
    for f in files:
        try:
            if os.path.exists(f):
                os.remove(f)
        except Exception:
            pass

def run_compare(df_old, df_new, selected_key_col_name):
    """Core comparison logic between two DataFrames."""
    if df_old.empty or df_new.empty:
        st.error("One of the files is empty or could not be read.")
        return False

    df_old = df_old.astype(str)
    df_new = df_new.astype(str)

    if df_old.columns.empty or df_new.columns.empty:
        st.error("No columns found in one of the files. Please check the file format.")
        return False

    common_cols = [col for col in df_old.columns if col in df_new.columns]
    if not common_cols:
        st.error("No common columns found between the two files.")
        return False

    display_key_col_name = selected_key_col_name
    compare_key_col_internal_name = "_compare_key_normalized"
    is_case_insensitive_key = False

    if display_key_col_name in df_old.columns and display_key_col_name in df_new.columns:
        st.info(f"Using '{display_key_col_name}' as the case-insensitive key column.")
        df_old[compare_key_col_internal_name] = df_old[display_key_col_name].str.upper()
        df_new[compare_key_col_internal_name] = df_new[display_key_col_name].str.upper()
        key_for_merge_and_grouping = compare_key_col_internal_name
        is_case_insensitive_key = True
    else:
        key_for_merge_and_grouping = df_old.columns[0]
        st.warning(
            f"Key column '{display_key_col_name}' was not found in both files. "
            f"Falling back to the first column '{key_for_merge_and_grouping}' (case-sensitive)."
        )
        df_old[compare_key_col_internal_name] = df_old[key_for_merge_and_grouping]
        df_new[compare_key_col_internal_name] = df_new[key_for_merge_and_grouping]

    df_old["_row_id"] = df_old.groupby(key_for_merge_and_grouping).cumcount()
    df_new["_row_id"] = df_new.groupby(key_for_merge_and_grouping).cumcount()

    df_compare = pd.merge(
        df_old,
        df_new,
        on=[key_for_merge_and_grouping, "_row_id"],
        how="outer",
        suffixes=("_old", "_new"),
        indicator=True,
    )

    df_compare["Status"] = ""
    df_compare["_changed_cols"] = ""

    ordered_original_cols = [
        col for col in df_old.columns if col not in ["_row_id", compare_key_col_internal_name]
    ]
    for col in df_new.columns:
        if col not in ["_row_id", compare_key_col_internal_name] and col not in ordered_original_cols:
            ordered_original_cols.append(col)

    actual_first_col_for_display = (
        display_key_col_name if is_case_insensitive_key else key_for_merge_and_grouping
    )
    if (
        actual_first_col_for_display in ordered_original_cols
        and ordered_original_cols[0] != actual_first_col_for_display
    ):
        ordered_original_cols.remove(actual_first_col_for_display)
        ordered_original_cols.insert(0, actual_first_col_for_display)
    elif actual_first_col_for_display not in ordered_original_cols:
        ordered_original_cols.insert(0, actual_first_col_for_display)

    temp_final_cols_data = {}
    for col in ordered_original_cols:
        if col != actual_first_col_for_display:
            temp_final_cols_data[col] = df_compare[f"{col}_new"].combine_first(df_compare[f"{col}_old"])

    temp_df_for_dupe_check = pd.DataFrame(temp_final_cols_data, index=df_compare.index)
    temp_df_for_dupe_check[key_for_merge_and_grouping] = df_compare[key_for_merge_and_grouping]

    dupes_mask_in_final = temp_df_for_dupe_check.duplicated(
        subset=[key_for_merge_and_grouping] + list(temp_final_cols_data.keys()),
        keep=False,
    )
    df_compare.loc[dupes_mask_in_final, "Status"] = "Duplicate"

    def get_changed_columns(row, original_cols_list, primary_key_display_name, is_pk_case_insensitive_flag):
        changed = []
        for col_base in original_cols_list:
            oldv = row.get(f"{col_base}_old", "")
            newv = row.get(f"{col_base}_new", "")
            if col_base == primary_key_display_name and is_pk_case_insensitive_flag:
                if str(oldv).upper() != str(newv).upper():
                    changed.append(col_base)
            else:
                if str(oldv) != str(newv):
                    changed.append(col_base)
        return changed

    for i, row in df_compare.iterrows():
        changed_cols = get_changed_columns(
            row,
            ordered_original_cols,
            actual_first_col_for_display,
            is_case_insensitive_key,
        )
        current_status_from_dupe_check = df_compare.at[i, "Status"]

        if row["_merge"] == "left_only":
            df_compare.at[i, "Status"] = "Deleted"

        elif row["_merge"] == "right_only":
            if current_status_from_dupe_check != "Duplicate":
                df_compare.at[i, "Status"] = "Added"

        else:
            if current_status_from_dupe_check != "Duplicate":
                if changed_cols:
                    df_compare.at[i, "Status"] = "Modified"
                else:
                    df_compare.at[i, "Status"] = "Same"

        df_compare.at[i, "_changed_cols"] = ",".join(changed_cols) if changed_cols else ""

    final_df = pd.DataFrame(index=df_compare.index)
    for col in ordered_original_cols:
        final_df[col] = df_compare[f"{col}_new"].combine_first(df_compare[f"{col}_old"])

    final_df["Status"] = df_compare["Status"]
    final_df["_changed_cols"] = df_compare["_changed_cols"]
    final_df = final_df.drop(columns=["_row_id"], errors="ignore")

    st.session_state["compare_df"] = final_df
    st.session_state["active_status_string"] = "All"
    st.session_state["select_all_rows"] = True
    st.success("Comparison completed.")
    return True

def merge_duplicates_action(df_current):
    """Merge duplicate rows, keeping the first row per key group."""
    dupes_mask = df_current["Status"] == "Duplicate"
    dupes = df_current[dupes_mask].copy()

    if dupes.empty:
        st.warning("No duplicate rows found to merge.")
        return

    key_col = df_current.columns[0]
    merged = dupes.groupby(key_col, as_index=False).first()
    merged["Status"] = "MergedDuplicate"

    remaining = df_current[~dupes_mask].copy()
    combined = pd.concat([remaining, merged], ignore_index=True)
    combined.index = pd.RangeIndex(start=0, stop=len(combined))

    st.session_state["compare_df"] = combined
    st.success(f"Duplicate rows merged. {len(dupes) - len(merged)} rows removed.")
    st.session_state["active_status_string"] = "Duplicate"

def handle_download(df_final, keep_indices, download_type):
    """Prepare download data and store it in session state."""
    st.session_state["download_data"] = None

    df_to_export = df_final[
        (df_final.index.isin(keep_indices)) | (df_final["Status"] == "Same")
    ].copy()

    if df_to_export.empty:
        st.warning("No data left to export after filtering.")
        return

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

    df_export_for_coloring = df_to_export.copy()
    df_export_final = df_to_export.drop(
        columns=["_changed_cols"],
        errors="ignore",
    ).copy()

    data_to_store = None

    if download_type == "plain":
        buffer = io.BytesIO()
        df_export_final.to_excel(buffer, index=False)
        buffer.seek(0)
        data_to_store = {
            "label": "Plain",
            "data": buffer.getvalue(),
            "filename": f"OUTPUT_PLAIN_{timestamp}.xlsx",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    elif download_type == "colored":
        file_colored = f"OUTPUT_COLORED_{timestamp}.xlsx"
        apply_excel_coloring(df_export_for_coloring, file_colored)
        excel_buffer = io.BytesIO(open(file_colored, "rb").read())
        excel_buffer.seek(0)
        cleanup_temp_files(file_colored)
        data_to_store = {
            "label": "Colored",
            "data": excel_buffer.getvalue(),
            "filename": f"OUTPUT_COLORED_{timestamp}.xlsx",
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    if data_to_store:
        st.session_state["download_data"] = data_to_store
        st.success(f"{data_to_store['label']} file is ready to download.")
        st.rerun()


def render_final_download_button():
    """Render the final download button if data is available."""
    data = st.session_state.get("download_data")
    if data:
        st.markdown("---")
        st.subheader("✅ Download Ready")
        st.download_button(
            label=f"Download {data['label']} File",
            data=data["data"],
            file_name=data["filename"],
            mime=data["mime"],
            key="final_download_trigger",
            use_container_width=True,
        )

# SESSION STATE INIT =========================

for key, val in [
    ("compare_df", None),
    ("active_status_string", "All"),
    ("select_all_rows", True),
    ("download_data", None),
    ("df_old_raw", None),
    ("df_new_raw", None),
    ("selected_key_col", None),
]:
    if key not in st.session_state:
        st.session_state[key] = val


# MAIN UI =========================

with st.container(border=True):
    st.subheader("1. Upload Files")

    col1, col2 = st.columns(2)
    with col1:
        file_old = st.file_uploader(
            "Old File (CSV/XLSX/XLSM)",
            type=["csv", "xlsx", "xlsm"],
            key="old_file",
        )
    with col2:
        file_new = st.file_uploader(
            "New File (CSV/XLSX/XLSM)",
            type=["csv", "xlsx", "xlsm"],
            key="new_file",
        )

    if st.button("Load Files & Select Key Column", use_container_width=True, type="primary"):
        if file_old and file_new:
            try:
                st.session_state["df_old_raw"] = read_file_to_df(file_old)
                st.session_state["df_new_raw"] = read_file_to_df(file_new)
                st.session_state["compare_df"] = None
                st.session_state["selected_key_col"] = None
                st.success("Files loaded successfully. Please select a key column.")
            except Exception as e:
                st.error(f"Error while loading files: {e}")
                traceback.print_exc()
        else:
            st.warning("Please upload both files before proceeding.")

if st.session_state["df_old_raw"] is not None and st.session_state["df_new_raw"] is not None:
    st.markdown("---")
    st.subheader("2. Select Key Column & Run Comparison")

    df_old_temp = st.session_state["df_old_raw"]
    df_new_temp = st.session_state["df_new_raw"]
    common_cols = [col for col in df_old_temp.columns if col in df_new_temp.columns]

    if not common_cols:
        st.error("No common columns found between the two files.")
    else:
        default_index = 0
        if st.session_state["selected_key_col"] in common_cols:
            default_index = common_cols.index(st.session_state["selected_key_col"])

        selected_key_col_name = st.selectbox(
            "Key column for comparison:",
            options=common_cols,
            index=default_index,
            key="key_col_selector_main",
        )
        st.session_state["selected_key_col"] = selected_key_col_name

        if st.button("Run Comparison", use_container_width=True, type="primary", key="run_comparison_button"):
            try:
                st.session_state["download_data"] = None
                run_compare(df_old_temp.copy(), df_new_temp.copy(), selected_key_col_name)
            except Exception as e:
                st.error(f"Error during comparison: {e}")
                traceback.print_exc()

if st.session_state["compare_df"] is not None:
    st.subheader("3. Comparison Preview (Non-Same Rows)")

    df_cached = st.session_state["compare_df"]
    df_preview_raw = df_cached[df_cached["Status"] != "Same"].copy()
    counts = get_compare_summary(df_cached)

    status_options = ["All", "Added", "Modified", "Deleted", "Duplicate"]
    tab_names = [
        f"All ({counts['All']})",
        f"Added (🟩 {counts['Added']})",
        f"Modified (🟨 {counts['Modified']})",
        f"Deleted (🟥 {counts['Deleted']})",
        f"Duplicate (🟦 {counts['Duplicate']})",
    ]

    current_status_string = st.session_state["active_status_string"]
    col_tab_buttons = st.columns(len(status_options))

    for i, col in enumerate(col_tab_buttons):
        is_active = status_options[i] == current_status_string
        with col:
            if st.button(
                tab_names[i],
                key=f"tab_btn_{i}",
                use_container_width=True,
                type="primary" if is_active else "secondary",
            ):
                st.session_state["active_status_string"] = status_options[i]
                st.session_state["select_all_rows"] = True
                st.session_state["download_data"] = None
                st.rerun()

    if current_status_string != "All":
        df_preview = df_preview_raw[df_preview_raw["Status"] == current_status_string].copy()
    else:
        df_preview = df_preview_raw.copy()
        status_order = {"Added": 1, "Modified": 2, "Deleted": 3, "Duplicate": 4, "MergedDuplicate": 4}
        df_preview["order"] = df_preview["Status"].map(status_order).fillna(99)
        df_preview = df_preview.sort_values(by="order").drop(columns=["order"])

    st.info(f"Displaying {len(df_preview)} rows with status: {current_status_string}")
    st.subheader("Data Preview")

    df_editor = df_preview.reset_index().rename(columns={"index": "RowIndex"})
    df_editor["RowIndex"] = df_editor["RowIndex"].astype(int)
    df_editor["Selected"] = st.session_state["select_all_rows"]

    status_emoji = {
        "Added": "🟩",
        "Deleted": "🟥",
        "Modified": "🟨",
        "Duplicate": "🟦",
        "MergedDuplicate": "🟦",
        "Same": "⬜",
    }
    df_editor["StatusBadge"] = df_editor["Status"].map(status_emoji).fillna("⬜")

    cols_to_exclude_from_preview = ["_row_id", "_compare_key_normalized", "_changed_cols"]
    display_cols_for_editor = [c for c in df_editor.columns if c not in cols_to_exclude_from_preview]

    col_order = ["StatusBadge", "Selected", "RowIndex"] + [
        c for c in display_cols_for_editor if c not in ["StatusBadge", "Selected", "RowIndex"]
    ]
    df_editor = df_editor[col_order]

    edited_df = st.data_editor(
        df_editor,
        column_config={
            "Selected": st.column_config.CheckboxColumn(
                "Selected",
                default=st.session_state["select_all_rows"],
            ),
            "RowIndex": st.column_config.Column("Row Index", width="small"),
            "StatusBadge": st.column_config.TextColumn(""),
        },
        hide_index=True,
        use_container_width=True,
        height=400,
        key="preview_editor",
    )

    keep_indices = edited_df.loc[edited_df["Selected"] == True, "RowIndex"].astype(int).tolist()
    st.markdown("---")

    if current_status_string == "Deleted":
        st.subheader("Bulk Action: Remove Deleted Rows")
        if st.button("Delete All Deleted Rows", key="btn_delete_all_deleted", type="primary"):
            df_new_cached = df_cached[df_cached["Status"] != "Deleted"].copy()
            st.session_state["compare_df"] = df_new_cached
            st.success("All rows with status 'Deleted' have been removed from the result.")
            st.rerun()

    if current_status_string == "Duplicate":
        st.subheader("Bulk Action: Merge Duplicate Rows")
        if st.button("Merge Duplicates", key="btn_merge", type="primary"):
            merge_duplicates_action(df_cached)
            st.rerun()

    if current_status_string == "All":
        st.subheader("Export Final Result")
        col_plain, col_colored = st.columns(2)

        with col_plain:
            if st.button(
                "Download Plain",
                type="primary",
                key="btn_download_plain",
                use_container_width=True,
            ):
                handle_download(df_cached, keep_indices, "plain")

        with col_colored:
            if st.button(
                "Download Colored",
                type="primary",
                key="btn_download_colored",
                use_container_width=True,
            ):
                handle_download(df_cached, keep_indices, "colored")
    else:
        st.markdown("Export options are available in the *All* tab only.")

    render_final_download_button()